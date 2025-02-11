import requests
import time
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference


API_URL = "https://api.coingecko.com/api/v3/coins/markets"


def fetch_crypto_data(limit=50):
    parameters = {
        "vs_currency": "usd",
        "order": "market_cap_desc",
        "per_page": limit,
        "page": 1,
        "sparkline": False
    }
    response = requests.get(API_URL, params=parameters)
    response.raise_for_status()
    return response.json()


def analyze_data(data):
    df = pd.DataFrame(data)

    top_5_by_market_cap = df.nlargest(5, 'market_cap')
    average_price = df['current_price'].mean()
    highest_24h_change = df.loc[df['price_change_percentage_24h'].idxmax()]
    lowest_24h_change = df.loc[df['price_change_percentage_24h'].idxmin()]

    return top_5_by_market_cap, average_price, highest_24h_change, lowest_24h_change, df


def update_excel(df, filename="crypto_data.xlsx"):
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

        sheet.delete_rows(2, sheet.max_row)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        header = ['Name', 'Symbol', 'Price', 'Market Cap', '24h Volume', '24h Change']
        sheet.append(header)

    for row in dataframe_to_rows(
            df[['name', 'symbol', 'current_price', 'market_cap', 'total_volume', 'price_change_percentage_24h']],
            index=False, header=False):
        sheet.append(row)


    chart = LineChart()
    chart.title = "Crypto Price Trend"
    chart.x_axis.title = "Index"
    chart.y_axis.title = "Price"
    data = Reference(sheet, min_col=3, min_row=2, max_col=3, max_row=sheet.max_row)
    chart.add_data(data)
    sheet.add_chart(chart, "G2")

    workbook.save(filename)


def generate_report(top_5, avg_price, highest_change, lowest_change, filename="analysis_report.txt"):
    with open(filename, "w") as f:
        f.write("Crypto Data Analysis Report\n\n")
        f.write("Top 5 Cryptocurrencies by Market Cap:\n")
        f.write(top_5.to_string() + "\n\n")
        f.write(f"Average Price: ${avg_price:.2f}\n\n")
        f.write("Highest 24h Price Change:\n")
        f.write(highest_change.to_string() + "\n\n")
        f.write("Lowest 24h Price Change:\n")
        f.write(lowest_change.to_string() + "\n")


if __name__ == "__main__":
    while True:
        try:
            crypto_data = fetch_crypto_data()
            top_5, avg_price, highest_change, lowest_change, df = analyze_data(crypto_data)
            update_excel(df)
            generate_report(top_5, avg_price, highest_change, lowest_change)
            print("Data updated. Waiting for next update...")
            time.sleep(300)
        except requests.exceptions.RequestException as e:
            print(f"Error fetching data: {e}")
            time.sleep(300)
        except Exception as e:
            print(f"An error occurred: {e}")
            time.sleep(300)